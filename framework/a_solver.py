"""Academic kinematic baseline for the uploaded 2025 modeling A problem.

This solver is deliberately deterministic and auditable. It models missile flight,
horizontal constant-speed UAV motion, ballistic release, cloud sinking, and a
10 m line-of-sight proxy around the true-target center. It is a baseline, not a
claim of global optimality or a real-world operational plan.
"""
from __future__ import annotations

import json
import math
import os
import shutil
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

G = 9.8
CLOUD_RADIUS = 10.0
CLOUD_LIFE = 20.0
CLOUD_SINK = 3.0
MISSILE_SPEED = 300.0
TARGET = (0.0, 200.0, 5.0)
MISSILES = {"M1": (20000.0, 0.0, 2000.0), "M2": (19000.0, 600.0, 2100.0), "M3": (18000.0, -600.0, 1900.0)}
UAVS = {"FY1": (17800.0, 0.0, 1800.0), "FY2": (12000.0, 1400.0, 1400.0), "FY3": (6000.0, -3000.0, 700.0), "FY4": (11000.0, 2000.0, 1800.0), "FY5": (13000.0, -2000.0, 1300.0)}

ROOT = Path(os.environ.get("FRAMEWORK_INPUT_DIR", "."))
OUT = Path(os.environ.get("FRAMEWORK_OUTPUT_DIR", ROOT / "outputs"))
OUT.mkdir(parents=True, exist_ok=True)

def sub(a, b): return tuple(x - y for x, y in zip(a, b))
def add(a, b): return tuple(x + y for x, y in zip(a, b))
def mul(a, k): return tuple(x * k for x in a)
def norm(a): return math.sqrt(sum(x * x for x in a))
def missile_time(name): return norm(MISSILES[name]) / MISSILE_SPEED
def missile_pos(name, t):
    start = MISSILES[name]
    total = missile_time(name)
    return mul(start, max(0.0, 1.0 - t / total))
def uav_position(name, theta, speed, t):
    start = UAVS[name]
    return (start[0] + speed * math.cos(theta) * t, start[1] + speed * math.sin(theta) * t, start[2])
def bomb_explosion(uav, theta, speed, release, fuse):
    p = uav_position(uav, theta, speed, release)
    dt = fuse
    return (p[0] + speed * math.cos(theta) * dt, p[1] + speed * math.sin(theta) * dt, p[2] - 0.5 * G * dt * dt)
def point_segment_distance(point, a, b):
    ab = sub(b, a)
    denom = sum(x * x for x in ab)
    if denom == 0: return norm(sub(point, a))
    t = max(0.0, min(1.0, sum((point[i] - a[i]) * ab[i] for i in range(3)) / denom))
    return norm(sub(point, add(a, mul(ab, t))))
def is_covered(missile, explosion, t, fuse):
    det = fuse
    if t < det or t > det + CLOUD_LIFE: return False
    cloud = add(explosion, (0.0, 0.0, -CLOUD_SINK * (t - det)))
    return point_segment_distance(cloud, missile_pos(missile, t), TARGET) <= CLOUD_RADIUS
def duration(missile, explosion, fuse, step=0.2):
    end = missile_time(missile)
    hits = [is_covered(missile, explosion, i * step, fuse) for i in range(int(end / step) + 1)]
    return sum(step for hit in hits if hit)
def candidate(uav, missile, theta, speed, release, fuse):
    explosion = bomb_explosion(uav, theta, speed, release, fuse)
    return {"uav": uav, "missile": missile, "theta": theta, "speed": speed, "release": release, "fuse": fuse, "release_point": uav_position(uav, theta, speed, release), "explosion_point": explosion, "duration": duration(missile, explosion, fuse)}
def best_single(uav, missile, coarse=True):
    end = min(45.0, missile_time(missile) - 2.0)
    theta_step = 20 if coarse else 10
    release_step = 4.0 if coarse else 2.0
    fuse_values = (2.0, 5.0, 8.0, 11.0, 14.0) if coarse else (1.0, 3.0, 5.0, 7.0, 9.0, 11.0, 13.0, 15.0)
    best = None
    for deg in range(0, 360, theta_step):
        theta = math.radians(deg)
        for speed in range(70, 141, 10):
            for release_i in range(int(end / release_step) + 1):
                release = release_i * release_step
                for fuse in fuse_values:
                    item = candidate(uav, missile, theta, float(speed), release, fuse)
                    if best is None or item["duration"] > best["duration"]: best = item
    return best
def fixed_q1():
    theta = math.pi
    item = candidate("FY1", "M1", theta, 120.0, 1.5, 3.6)
    item["question"] = "问题1固定条件"
    return item
def copy_template(name):
    candidates = list(ROOT.rglob(name)) if ROOT.exists() else []
    source = candidates[0] if candidates else ROOT / name
    target = OUT / name
    if source.exists(): shutil.copy2(source, target)
    elif load_workbook: load_workbook().save(target)
    return target
def put(ws, row, values):
    for col, value in enumerate(values, 1): ws.cell(row=row, column=col, value=value)
def export_outputs(q2):
    if not load_workbook:
        return ["openpyxl unavailable; xlsx outputs were not written"]
    outputs = []
    wb = load_workbook(copy_template("result1.xlsx")); ws = wb.active
    q3 = [candidate("FY1", "M1", q2["theta"] + offset, q2["speed"], q2["release"] + i * 3.0, q2["fuse"]) for i, offset in enumerate((0.0, 0.08, -0.08))]
    for i, item in enumerate(q3, 2): put(ws, i, [math.degrees(item["theta"]) % 360, item["speed"], i - 1, *item["release_point"], *item["explosion_point"], item["duration"]])
    wb.save(OUT / "result1.xlsx"); outputs.append("result1.xlsx")
    wb = load_workbook(copy_template("result2.xlsx")); ws = wb.active
    for i, uav in enumerate(("FY1", "FY2", "FY3"), 2):
        item = best_single(uav, "M1", coarse=True); put(ws, i, [uav, math.degrees(item["theta"]) % 360, item["speed"], *item["release_point"], *item["explosion_point"], item["duration"]])
    wb.save(OUT / "result2.xlsx"); outputs.append("result2.xlsx")
    wb = load_workbook(copy_template("result3.xlsx")); ws = wb.active
    row = 2
    for uav in UAVS:
        target_missile = ("M1", "M2", "M3")[(row - 2) % 3]
        base = best_single(uav, target_missile, coarse=True)
        for bomb in range(1, 4):
            item = candidate(uav, target_missile, base["theta"] + (bomb - 2) * 0.05, base["speed"], base["release"] + (bomb - 1) * 3.0, base["fuse"])
            put(ws, row, [uav, math.degrees(item["theta"]) % 360, item["speed"], bomb, *item["release_point"], *item["explosion_point"], item["duration"], target_missile]); row += 1
    wb.save(OUT / "result3.xlsx"); outputs.append("result3.xlsx")
    return outputs

q1 = fixed_q1()
q2 = best_single("FY1", "M1", coarse=bool(os.environ.get("FRAMEWORK_FAST_TEST")))
outputs = export_outputs(q2)
summary = {"solver": "deterministic-kinematic-baseline-v1", "q1_fixed_duration_seconds": q1["duration"], "q2_best_duration_seconds": q2["duration"], "outputs": outputs, "warning": "结果是竞赛建模启发式基线，需人工复核模型假设、目标函数和模板填写，不代表全局最优。"}
(OUT / "solver_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps(summary, ensure_ascii=False))
